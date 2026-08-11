#!/usr/bin/env python3
"""Build MONEY 2026 — Dashboard chính + Job Agency + multi-job (LG-style) + link Bluescope."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent / "MONEY_2026_System.xlsx"

# External Bluescope booking (keep as separate spreadsheet)
BLUESCOPE_ID = "1ygn6O9Ej1-AmaeCVDOp6UD3pnbiyhG0c-ZUJ4MgNzdo"
BLUESCOPE_URL = (
    f"https://docs.google.com/spreadsheets/d/{BLUESCOPE_ID}/edit"
)
# Main MONEY target (existing file to replace content / upload into)
MONEY_ID = "12dpkDrLw1Cban2KNyNiHJ6vfqun_Las8Qa_ZI1cA3Js"
MONEY_URL = f"https://docs.google.com/spreadsheets/d/{MONEY_ID}/edit"

# Styles
THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
FILL_HEADER = PatternFill("solid", fgColor="0F172A")
FILL_SECTION = PatternFill("solid", fgColor="1E293B")
FILL_INPUT = PatternFill("solid", fgColor="FEF3C7")  # yellow = nhập liệu
FILL_CALC = PatternFill("solid", fgColor="E0F2FE")  # blue = tự tính
FILL_OK = PatternFill("solid", fgColor="DCFCE7")
FILL_WARN = PatternFill("solid", fgColor="FEE2E2")
FILL_SOFT = PatternFill("solid", fgColor="F8FAFC")
FILL_LINK = PatternFill("solid", fgColor="EDE9FE")
FILL_GREEN = PatternFill("solid", fgColor="BBF7D0")
FILL_ORANGE = PatternFill("solid", fgColor="FFEDD5")

FONT_WHITE = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(name="Arial", bold=True, size=16, color="0F172A")
FONT_SECTION = Font(name="Arial", bold=True, size=12, color="FFFFFF")
FONT_BOLD = Font(name="Arial", bold=True, size=11)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_SMALL = Font(name="Arial", size=9, color="64748B")
FONT_MONEY = Font(name="Arial", bold=True, size=12, color="0F172A")
FONT_LINK = Font(name="Arial", size=10, color="4F46E5", underline="single")

VND = '#,##0'
PCT = '0%'


def style_header_row(ws, row, start_col, end_col, fill=FILL_HEADER):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = FONT_WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def style_input(cell):
    cell.fill = FILL_INPUT
    cell.border = THIN
    cell.font = FONT_NORMAL


def style_calc(cell):
    cell.fill = FILL_CALC
    cell.border = THIN
    cell.font = FONT_NORMAL


def money(cell, is_input=False):
    cell.number_format = VND
    if is_input:
        style_input(cell)
    else:
        style_calc(cell)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_title(ws, text, row=1, merge_to="H"):
    ws.merge_cells(f"A{row}:{merge_to}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_TITLE
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 28


def build_huong_dan(wb):
    ws = wb.create_sheet("00_HUONG_DAN", 0)
    set_col_widths(ws, [4, 90])
    add_title(ws, "MONEY 2026 — Hướng dẫn hệ thống", 1, "B")

    lines = [
        "",
        "MỤC TIÊU",
        "• 1 Dashboard chính nhìn toàn bộ tiền / nợ / công nợ / P&L theo job",
        "• Job Agency: danh sách mọi job (Agency + Job), mỗi job 1 sheet riêng kiểu LG",
        "• Bluescope Booking giữ FILE RIÊNG, Dashboard chỉ kéo TỔNG CÔNG NỢ / ngân sách",
        "• Thêm job mới = nhân bản sheet _TEMPLATE_JOB → đổi tên JOB_XXX → gắn vào 02_JOB_AGENCY",
        "",
        "CẤU TRÚC FILE",
        "• 01_DASHBOARD      — Dashboard chính (KPI + tổng công nợ + tổng job)",
        "• 02_JOB_AGENCY     — Sổ cái job/agency (registry + link sheet + link ngoài)",
        "• 03_CONG_NO        — Công nợ khách hàng (tổng hợp, bao gồm Bluescope)",
        "• 04_NO_CA_NHAN     — Nợ cá nhân (MOM, Trí) + lịch trả",
        "• 05_TIN_DUNG       — Dư nợ thẻ tín dụng theo tháng",
        "• _TEMPLATE_JOB     — Mẫu chi phí job (copy để tạo job mới như LG)",
        "• JOB_LG            — Ví dụ job LG (đã seed)",
        "• JOB_BLUESCOPE     — Tóm tắt job Bluescope (link file ngoài)",
        "• 99_CAU_HINH       — Link sheet ngoài + % phí + % A Tân",
        "",
        "LINK BLUESCOPE (FILE RIÊNG — KHÔNG GỘP VÀO MONEY)",
        f"• URL: {BLUESCOPE_URL}",
        f"• ID:  {BLUESCOPE_ID}",
        "• Sheet gợi ý: BLUESCOPE_2025 (Rate card + booking event) | Ratecard video | Sheet5 content list",
        "• Trên Google Sheets, tại JOB_BLUESCOPE / 03_CONG_NO dùng IMPORTRANGE sau khi Allow access:",
        f'  =IMPORTRANGE("{BLUESCOPE_URL}";"BLUESCOPE_2025!L2")   ← Ngân sách',
        f'  =IMPORTRANGE("{BLUESCOPE_URL}";"BLUESCOPE_2025!N2")   ← Đã xài',
        f'  =IMPORTRANGE("{BLUESCOPE_URL}";"BLUESCOPE_2025!O2")   ← Còn lại',
        f'  =SUM(IMPORTRANGE("{BLUESCOPE_URL}";"BLUESCOPE_2025!N9:N100"))  ← Tổng chi event',
        "",
        "CÁCH THÊM JOB MỚI (KIỂU LG)",
        "1. Mở sheet _TEMPLATE_JOB → chuột phải tab → Duplicate",
        "2. Đổi tên tab thành JOB_<TEN>  (vd: JOB_SUNGROUP, JOB_PBCM) — không dấu, không khoảng",
        "3. Điền khối A. HỢP ĐỒNG (vàng) + B. CHI PHÍ (vàng)",
        "4. Vào 02_JOB_AGENCY thêm 1 dòng: Agency | Tên job | Tên sheet | Link (nếu có)",
        "5. Các cột Tổng HĐ / Chi phí / LN / Còn lại sẽ tự lấy từ sheet job (công thức sẵn)",
        "6. Dashboard tự cộng dòng mới nếu bạn mở rộng vùng SUM (xem ghi chú trên Dashboard)",
        "",
        "MÀU Ô",
        "• Vàng nhạt  = nhập tay",
        "• Xanh dương nhạt = công thức tự tính — không sửa",
        "• Tím nhạt = link / sheet ngoài",
        "",
        "UPLOAD LÊN GOOGLE SHEETS",
        "1. Upload file MONEY_2026_System.xlsx lên Drive (DOC / MONEY 2026)",
        "2. Mở → File → Save as Google Sheets (hoặc Open with Google Sheets)",
        f"3. Có thể dán nội dung vào file hiện tại: {MONEY_URL}",
        "4. Cho phép IMPORTRANGE khi sheet hỏi Allow access (1 lần)",
        "5. Bluescope vẫn mở bằng link riêng — không copy booking vào Money",
        "",
        "QUY ƯỚC TÍNH (chỉnh ở 99_CAU_HINH)",
        "• Phí xuất HĐ mặc định 4% trên Tổng HĐ",
        "• A Tân mặc định 20% trên Lợi nhuận gộp (HĐ − chi phí)",
        "• Còn lại công nợ = Tổng HĐ − Đã thu",
    ]
    for i, line in enumerate(lines, 2):
        cell = ws.cell(row=i, column=2, value=line)
        if line.startswith("•") or line.startswith("  =") or line.startswith("1.") or line.startswith("2.") or line.startswith("3.") or line.startswith("4.") or line.startswith("5.") or line.startswith("6."):
            cell.font = FONT_NORMAL
        elif line and not line.startswith(" "):
            cell.font = FONT_BOLD
        else:
            cell.font = FONT_SMALL
    ws.freeze_panes = "A3"


def build_cau_hinh(wb):
    ws = wb.create_sheet("99_CAU_HINH", 1)
    set_col_widths(ws, [4, 36, 55, 40])
    add_title(ws, "99 · CẤU HÌNH HỆ THỐNG", 1, "D")

    ws["B3"] = "Tham số"
    ws["C3"] = "Giá trị"
    ws["D3"] = "Ghi chú"
    style_header_row(ws, 3, 2, 4)

    rows = [
        ("Phi_Xuat_HD", 0.04, "Phí xuất hoá đơn (4%) — dùng trong job template"),
        ("Ty_Le_A_Tan", 0.20, "A Tân nhận % lợi nhuận gộp"),
        ("Tien_Hien_Co", 71_000_000, "Cập nhật tay — Dashboard đọc ô này"),
        ("No_Goc_MOM", 88_000_000, "Nợ gốc MOM"),
        ("Tra_Moi_Thang_MOM", 2_000_000, "Trả mỗi tháng MOM"),
        ("No_Goc_Tri", 175_000_000, "Nợ gốc a Trí"),
        ("Bluescope_Sheet_ID", BLUESCOPE_ID, "ID file booking Bluescope (riêng)"),
        ("Bluescope_URL", BLUESCOPE_URL, "Link mở Bluescope"),
        ("Bluescope_Tab", "BLUESCOPE_2025", "Tên tab chính trong file Bluescope"),
        ("Money_Sheet_ID", MONEY_ID, "ID file MONEY 2026 chính"),
    ]
    for i, (k, v, note) in enumerate(rows, 4):
        ws.cell(row=i, column=2, value=k).font = FONT_BOLD
        c = ws.cell(row=i, column=3, value=v)
        style_input(c)
        if isinstance(v, float) and v < 1:
            c.number_format = PCT
        elif isinstance(v, (int, float)):
            c.number_format = VND
        ws.cell(row=i, column=4, value=note).font = FONT_SMALL

    ws["B15"] = "Công thức IMPORTRANGE mẫu (dán vào Google Sheets)"
    ws["B15"].font = FONT_BOLD
    ws.merge_cells("B16:D16")
    ws["B16"] = (
        f'=IMPORTRANGE("{BLUESCOPE_URL}";"BLUESCOPE_2025!L2")'
    )
    ws["B16"].fill = FILL_LINK
    ws["B17"] = "Lần đầu Google Sheets sẽ hiện Allow access — bấm Allow."
    ws["B17"].font = FONT_SMALL


def build_job_template(wb, name="_TEMPLATE_JOB", seed=None):
    """seed: dict optional with title, agency, contract, costs..."""
    ws = wb.create_sheet(name)
    set_col_widths(ws, [4, 32, 18, 16, 36, 18])

    title = (seed or {}).get("title", "🎬 CHI PHÍ JOB — (đổi tên sheet thành JOB_XXX)")
    agency = (seed or {}).get("agency", "")
    job_name = (seed or {}).get("job_name", "")
    add_title(ws, title, 1, "F")

    ws["B2"] = "Agency / Client"
    ws["C2"] = agency
    style_input(ws["C2"])
    ws["D2"] = "Tên job"
    ws["E2"] = job_name
    style_input(ws["E2"])
    ws.merge_cells("E2:F2")

    ws["B3"] = "Trạng thái"
    ws["C3"] = (seed or {}).get("status", "Đang làm")
    style_input(ws["C3"])
    ws["D3"] = "Link file ngoài (nếu có)"
    ws["E3"] = (seed or {}).get("link", "")
    style_input(ws["E3"])
    ws.merge_cells("E3:F3")

    # A. HĐ
    ws.merge_cells("B5:F5")
    ws["B5"] = "A. HỢP ĐỒNG"
    ws["B5"].fill = FILL_SECTION
    ws["B5"].font = FONT_SECTION

    ws["B6"] = "Tổng phí ký hợp đồng"
    c = ws["C6"]
    c.value = (seed or {}).get("contract", 0)
    money(c, is_input=True)
    ws["D6"] = "← Nhập tổng giá trị HĐ"
    ws["D6"].font = FONT_SMALL

    ws["B7"] = "Phí xuất hoá đơn"
    c = ws["C7"]
    # reference 99_CAU_HINH!C4 for rate
    c.value = f"=C6*'99_CAU_HINH'!C4"
    money(c)
    ws["D7"] = "=% phí × Tổng HĐ (cấu hình)"
    ws["D7"].font = FONT_SMALL

    ws["B8"] = "Đã thu từ KH"
    c = ws["C8"]
    c.value = (seed or {}).get("collected", 0)
    money(c, is_input=True)
    ws["D8"] = "← Số đã nhận"
    ws["D8"].font = FONT_SMALL

    ws["B9"] = "Còn phải thu (công nợ)"
    c = ws["C9"]
    c.value = "=C6-C8"
    money(c)
    ws["D9"] = "Tổng HĐ − Đã thu"
    ws["D9"].font = FONT_SMALL

    # B. Chi phí
    ws.merge_cells("B11:F11")
    ws["B11"] = "B. CHI PHÍ SẢN XUẤT (nhập từng dòng)"
    ws["B11"].fill = FILL_SECTION
    ws["B11"].font = FONT_SECTION

    headers = ["#", "Hạng mục", "Số tiền (đ)", "Ghi chú", "Người nhận", "Đã chi?"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=12, column=i, value=h)
    style_header_row(ws, 12, 1, 6)

    default_items = (seed or {}).get(
        "items",
        [
            ("Camop 1", 0, "Camera Operator 1", "", "Chưa"),
            ("Camop 2", 0, "Camera Operator 2", "", "Chưa"),
            ("Producer", 0, "Producer / Đạo diễn", "", "Chưa"),
            ("Catering", 0, "Ăn uống / hậu cần", "", "Chưa"),
            ("Talent", 0, "Diễn viên / Người mẫu", "", "Chưa"),
            ("Edit / Post", 0, "Dựng / post", "", "Chưa"),
            ("Thiết bị / khác", 0, "Thuê máy, location…", "", "Chưa"),
            ("Chi khác 1", 0, "", "", "Chưa"),
            ("Chi khác 2", 0, "", "", "Chưa"),
            ("Chi khác 3", 0, "", "", "Chưa"),
        ],
    )

    start = 13
    for i, item in enumerate(default_items):
        r = start + i
        name_i, amt, note, person, paid = item if len(item) == 5 else (*item, "", "Chưa")
        ws.cell(row=r, column=1, value=i + 1).border = THIN
        b = ws.cell(row=r, column=2, value=name_i)
        style_input(b)
        c = ws.cell(row=r, column=3, value=amt)
        money(c, is_input=True)
        d = ws.cell(row=r, column=4, value=note)
        style_input(d)
        e = ws.cell(row=r, column=5, value=person)
        style_input(e)
        f = ws.cell(row=r, column=6, value=paid)
        style_input(f)

    end = start + len(default_items) - 1
    # extra blank rows for expansion
    for i in range(5):
        r = end + 1 + i
        ws.cell(row=r, column=1, value=len(default_items) + 1 + i).border = THIN
        for col in range(2, 7):
            style_input(ws.cell(row=r, column=col, value="" if col != 3 else 0))
            if col == 3:
                money(ws.cell(row=r, column=col), is_input=True)
    end_all = end + 5

    total_row = end_all + 2
    ws.cell(row=total_row, column=2, value="Tổng chi phí sản xuất").font = FONT_BOLD
    c = ws.cell(row=total_row, column=3, value=f"=SUM(C{start}:C{end_all})")
    money(c)

    # C. Tổng kết
    sum_start = total_row + 2
    ws.merge_cells(f"B{sum_start}:F{sum_start}")
    ws.cell(row=sum_start, column=2, value="C. TỔNG KẾT P&L").fill = FILL_SECTION
    ws.cell(row=sum_start, column=2).font = FONT_SECTION

    rows_sum = [
        (sum_start + 1, "Tổng chi phí (HĐ fee + SX)", f"=C7+C{total_row}", "Phí HĐ + SX"),
        (sum_start + 2, "Lợi nhuận gộp", f"=C6-C{sum_start+1}", "Tổng HĐ − Tổng chi phí"),
        (sum_start + 3, "💰 A Tân (share)", f"=C{sum_start+2}*'99_CAU_HINH'!C5", "% LN từ cấu hình"),
        (sum_start + 4, "Lợi nhuận còn lại", f"=C{sum_start+2}-C{sum_start+3}", "Sau A Tân"),
        (sum_start + 5, "Công nợ còn lại", "=C9", "Link sang 03_CONG_NO / Agency"),
    ]
    for r, label, formula, note in rows_sum:
        ws.cell(row=r, column=2, value=label).font = FONT_BOLD
        c = ws.cell(row=r, column=3, value=formula)
        money(c)
        ws.cell(row=r, column=4, value=note).font = FONT_SMALL

    # Named export cells for Agency (fixed positions for INDEX/INDIRECT)
    # We put key metrics in a fixed block at column H for easy lookup
    ws["H1"] = "EXPORT_KPI"
    ws["H1"].font = FONT_BOLD
    ws["H2"] = "Agency"
    ws["I2"] = "=C2"
    ws["H3"] = "Job"
    ws["I3"] = "=E2"
    ws["H4"] = "Tong_HD"
    ws["I4"] = "=C6"
    money(ws["I4"])
    ws["H5"] = "Da_Thu"
    ws["I5"] = "=C8"
    money(ws["I5"])
    ws["H6"] = "Con_Lai"
    ws["I6"] = "=C9"
    money(ws["I6"])
    ws["H7"] = "Chi_Phi"
    ws["I7"] = f"=C{sum_start+1}"
    money(ws["I7"])
    ws["H8"] = "Loi_Nhuan"
    ws["I8"] = f"=C{sum_start+2}"
    money(ws["I8"])
    ws["H9"] = "A_Tan"
    ws["I9"] = f"=C{sum_start+3}"
    money(ws["I9"])
    ws["H10"] = "LN_Con_Lai"
    ws["I10"] = f"=C{sum_start+4}"
    money(ws["I10"])
    ws["H11"] = "Trang_Thai"
    ws["I11"] = "=C3"
    ws["H12"] = "Link"
    ws["I12"] = "=E3"

    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16

    ws["B" + str(sum_start + 8)] = (
        "📌 Sau khi copy sheet: đổi tên tab JOB_XXX → cập nhật dòng tương ứng ở 02_JOB_AGENCY (cột Tên sheet)."
    )
    ws["B" + str(sum_start + 8)].font = FONT_SMALL

    # Data validation status
    dv = DataValidation(
        type="list",
        formula1='"Đang làm,Chờ thu,Đã xong,Tạm dừng"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(ws["C3"])

    dv2 = DataValidation(type="list", formula1='"Chưa,Một phần,Xong"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add(f"F{start}:F{end_all}")

    return ws


def build_job_lg(wb):
    seed = {
        "title": "🎬 JOB_LG — Production (mẫu job kiểu LG)",
        "agency": "LG",
        "job_name": "LG Production",
        "status": "Chờ thu",
        "contract": 63_104_000,
        "collected": 0,
        "items": [
            ("Camop 1", 8_000_000, "Camera Operator 1", "", "Chưa"),
            ("Camop 2", 6_000_000, "Camera Operator 2", "", "Chưa"),
            ("Producer", 5_000_000, "Producer", "", "Chưa"),
            ("Catering", 2_000_000, "Ăn uống", "", "Chưa"),
            ("Talent", 3_000_000, "Talent", "", "Chưa"),
            ("Edit / Post", 10_000_000, "Dựng", "", "Chưa"),
            ("Thiết bị / khác", 4_000_000, "Thuê TB", "", "Chưa"),
            ("Chi khác 1", 0, "", "", "Chưa"),
            ("Chi khác 2", 0, "", "", "Chưa"),
            ("Chi khác 3", 0, "", "", "Chưa"),
        ],
    }
    return build_job_template(wb, "JOB_LG", seed)


def build_job_bluescope(wb):
    ws = wb.create_sheet("JOB_BLUESCOPE")
    set_col_widths(ws, [4, 34, 20, 18, 50])
    add_title(ws, "🎬 JOB_BLUESCOPE — Booking riêng (link file ngoài)", 1, "E")

    ws.merge_cells("B2:E2")
    ws["B2"] = (
        f"⚡ File booking RIÊNG — không gộp vào Money. Mở: {BLUESCOPE_URL}"
    )
    ws["B2"].fill = FILL_LINK
    ws["B2"].font = FONT_LINK
    ws["B2"].hyperlink = BLUESCOPE_URL

    ws["B4"] = "Agency"
    ws["C4"] = "Bluescope / NS BlueScope"
    style_input(ws["C4"])
    ws["B5"] = "Tên job"
    ws["C5"] = "Booking production 2025–2026"
    style_input(ws["C5"])
    ws["B6"] = "Trạng thái"
    ws["C6"] = "Đang làm"
    style_input(ws["C6"])

    ws.merge_cells("B8:E8")
    ws["B8"] = "A. TỔNG HỢP TỪ FILE BOOKING (Google Sheets: thay bằng IMPORTRANGE)"
    ws["B8"].fill = FILL_SECTION
    ws["B8"].font = FONT_SECTION

    ws["B9"] = "Ngân sách (từ Bluescope L2)"
    c = ws["C9"]
    c.value = 100_000_000  # seed from public export; on GS use IMPORTRANGE
    money(c, is_input=True)
    ws["D9"] = "Excel: nhập tay / GS: IMPORTRANGE L2"
    ws["D9"].font = FONT_SMALL

    ws["B10"] = "Đã nhận"
    c = ws["C10"]
    c.value = 0
    money(c, is_input=True)

    ws["B11"] = "Đã xài (từ Bluescope N2)"
    c = ws["C11"]
    c.value = 44_000_000
    money(c, is_input=True)
    ws["D11"] = "Hoặc SUM cột TỔNG event"
    ws["D11"].font = FONT_SMALL

    ws["B12"] = "Còn lại ngân sách"
    c = ws["C12"]
    c.value = "=C9-C11"
    money(c)

    ws["B13"] = "Tổng HĐ / giá trị booking (dùng cho công nợ)"
    c = ws["C13"]
    c.value = "=C9"  # treat budget as contract value for receivable tracking
    money(c)

    ws["B14"] = "Đã thu từ KH"
    c = ws["C14"]
    c.value = 0
    money(c, is_input=True)

    ws["B15"] = "Công nợ còn lại"
    c = ws["C15"]
    c.value = "=C13-C14"
    money(c)

    ws.merge_cells("B17:E17")
    ws["B17"] = "B. SỰ KIỆN ĐÃ BOOK (tóm tắt — chi tiết ở file Bluescope)"
    ws["B17"].fill = FILL_SECTION
    ws["B17"].font = FONT_SECTION

    headers = ["#", "Tên sự kiện", "Ngày", "Tổng (đ)", "Ghi chú"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=18, column=i, value=h)
    style_header_row(ws, 18, 1, 5)

    events = [
        (1, "Quay Trồng Cây", "2026-06-11", 13_000_000, "Đi tỉnh"),
        (2, "Quay Bình Dương", "2026-06-17", 16_000_000, "Đi xa + NS"),
        (3, "Sửa POSM", "2026-06-26", 3_000_000, ""),
        (4, "Recap Q3", "2025-06-06", 8_000_000, "Edit"),
        (5, "POSM 2 banner", "2025-07-04", 2_000_000, ""),
        (6, "Recap Q3 fix 2", "2026-07-19", 2_000_000, "HCM"),
    ]
    for i, (stt, ten, ngay, tong, note) in enumerate(events):
        r = 19 + i
        ws.cell(row=r, column=1, value=stt).border = THIN
        style_input(ws.cell(row=r, column=2, value=ten))
        style_input(ws.cell(row=r, column=3, value=ngay))
        c = ws.cell(row=r, column=4, value=tong)
        money(c, is_input=True)
        style_input(ws.cell(row=r, column=5, value=note))

    for i in range(6, 16):
        r = 19 + i
        ws.cell(row=r, column=1, value=i + 1).border = THIN
        for col in range(2, 6):
            if col == 4:
                money(ws.cell(row=r, column=col, value=0), is_input=True)
            else:
                style_input(ws.cell(row=r, column=col, value=""))

    ws["B35"] = "Tổng chi event (sheet này)"
    c = ws["C35"]
    c.value = "=SUM(D19:D34)"
    money(c)

    # EXPORT block (same layout as template for Agency formulas)
    ws["H1"] = "EXPORT_KPI"
    ws["H1"].font = FONT_BOLD
    mapping = [
        (2, "Agency", "=C4"),
        (3, "Job", "=C5"),
        (4, "Tong_HD", "=C13"),
        (5, "Da_Thu", "=C14"),
        (6, "Con_Lai", "=C15"),
        (7, "Chi_Phi", "=C11"),
        (8, "Loi_Nhuan", "=C13-C11"),
        (9, "A_Tan", 0),
        (10, "LN_Con_Lai", "=I8-I9"),
        (11, "Trang_Thai", "=C6"),
        (12, "Link", BLUESCOPE_URL),
    ]
    for r, label, val in mapping:
        ws.cell(row=r, column=8, value=label)
        cell = ws.cell(row=r, column=9, value=val)
        if r in (4, 5, 6, 7, 8, 9, 10):
            money(cell)
        else:
            style_calc(cell) if r != 12 else style_input(cell)

    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 18

    ws["B37"] = "Công thức Google Sheets gợi ý (dán đè C9 / C11):"
    ws["B37"].font = FONT_BOLD
    ws["B38"] = f'=IMPORTRANGE("{BLUESCOPE_URL}";"BLUESCOPE_2025!L2")'
    ws["B38"].fill = FILL_LINK
    ws["B39"] = f'=IMPORTRANGE("{BLUESCOPE_URL}";"BLUESCOPE_2025!N2")'
    ws["B39"].fill = FILL_LINK
    ws["B40"] = (
        "Sau khi Allow access, Còn lại / công nợ / Dashboard sẽ cập nhật theo file Bluescope."
    )
    ws["B40"].font = FONT_SMALL


def build_job_agency(wb):
    ws = wb.create_sheet("02_JOB_AGENCY", 2)
    set_col_widths(ws, [5, 18, 22, 16, 14, 14, 14, 14, 14, 12, 40, 18])
    add_title(ws, "02 · JOB AGENCY — Sổ cái job (nhiều job kiểu LG)", 1, "L")

    ws.merge_cells("A2:L2")
    ws["A2"] = (
        "Mỗi dòng = 1 job. Cột 'Tên sheet' phải trùng tên tab (JOB_LG, JOB_BLUESCOPE, …). "
        "Các cột tiền tự lấy từ EXPORT_KPI (I4:I11) của sheet job. Thêm job: copy _TEMPLATE_JOB + thêm dòng."
    )
    ws["A2"].font = FONT_SMALL
    ws["A2"].fill = FILL_SOFT

    headers = [
        "#",
        "Agency",
        "Tên job",
        "Tên sheet",
        "Tổng HĐ",
        "Đã thu",
        "Còn lại",
        "Chi phí",
        "Lợi nhuận",
        "Trạng thái",
        "Link ngoài",
        "Ghi chú",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 1, 12)
    ws.row_dimensions[4].height = 22

    # Pre-seed jobs: LG + Bluescope + empty slots for more
    jobs = [
        (1, "LG", "LG Production", "JOB_LG", ""),
        (2, "Bluescope", "Booking 2025–2026", "JOB_BLUESCOPE", BLUESCOPE_URL),
        (3, "Bizeyes - Thảo Hiếu", "Rivus", "", ""),
        (4, "Bizeyes - Hiếu", "Masteris", "", ""),
        (5, "Bizeyes - Tuan", "Elyse", "", ""),
        (6, "2 Mic", "2 Mic", "", ""),
        (7, "a Trí", "Job a Trí", "", ""),
        (8, "Sky slois", "Sky slois", "", ""),
        (9, "Sun Group", "Sun Group", "", ""),
        (10, "pbcm", "pbcm", "", ""),
    ]

    # For jobs with sheet: use INDIRECT. For jobs without sheet: manual input on contract cols.
    # Excel INDIRECT with sheet name: INDIRECT("'"&D5&"'!I4")
    for idx, (stt, agency, job, sheet, link) in enumerate(jobs):
        r = 5 + idx
        ws.cell(row=r, column=1, value=stt).border = THIN
        style_input(ws.cell(row=r, column=2, value=agency))
        style_input(ws.cell(row=r, column=3, value=job))
        style_input(ws.cell(row=r, column=4, value=sheet))

        if sheet:
            # pull from job sheet export
            for col, cell_ref in [
                (5, "I4"),
                (6, "I5"),
                (7, "I6"),
                (8, "I7"),
                (9, "I8"),
            ]:
                c = ws.cell(
                    row=r,
                    column=col,
                    value=f"=IF(D{r}=\"\",\"\",IFERROR(INDIRECT(\"'\"&D{r}&\"'!{cell_ref}\"),0))",
                )
                money(c)
            c = ws.cell(
                row=r,
                column=10,
                value=f"=IF(D{r}=\"\",\"\",IFERROR(INDIRECT(\"'\"&D{r}&\"'!I11\"),\"\"))",
            )
            style_calc(c)
            c = ws.cell(
                row=r,
                column=11,
                value=f"=IF(D{r}=\"\",\"\",IFERROR(INDIRECT(\"'\"&D{r}&\"'!I12\"),\"{link}\"))",
            )
            style_calc(c)
        else:
            # manual until sheet created
            for col, default in [(5, 0), (6, 0), (7, None), (8, 0), (9, 0)]:
                if col == 7:
                    c = ws.cell(row=r, column=col, value=f"=E{r}-F{r}")
                    money(c)
                else:
                    c = ws.cell(row=r, column=col, value=default)
                    money(c, is_input=True)
            style_input(ws.cell(row=r, column=10, value="Chưa gắn sheet"))
            style_input(ws.cell(row=r, column=11, value=link))

        style_input(ws.cell(row=r, column=12, value=""))

    # Extra empty rows for expansion (11–25)
    for i in range(11, 26):
        r = 4 + i  # rows 15..29? wait: idx 0 row 5, so i=11 -> row 15
        r = 5 + i - 1  # row 15 for i=11
        ws.cell(row=r, column=1, value=i).border = THIN
        for col in range(2, 5):
            style_input(ws.cell(row=r, column=col, value=""))
        # if sheet name filled, use INDIRECT; else manual
        for col, cell_ref in [
            (5, "I4"),
            (6, "I5"),
            (7, "I6"),
            (8, "I7"),
            (9, "I8"),
        ]:
            if col == 7:
                # còn lại: if has sheet pull, else E-F
                c = ws.cell(
                    row=r,
                    column=col,
                    value=(
                        f'=IF(D{r}<>"",IFERROR(INDIRECT("\'"&D{r}&"\'!I6"),0),E{r}-F{r})'
                    ),
                )
            else:
                c = ws.cell(
                    row=r,
                    column=col,
                    value=(
                        f'=IF(D{r}<>"",IFERROR(INDIRECT("\'"&D{r}&"\'!{cell_ref}"),0),0)'
                    ),
                )
            money(c)
        c = ws.cell(
            row=r,
            column=10,
            value=f'=IF(D{r}<>"",IFERROR(INDIRECT("\'"&D{r}&"\'!I11"),""),"")',
        )
        style_calc(c)
        c = ws.cell(
            row=r,
            column=11,
            value=f'=IF(D{r}<>"",IFERROR(INDIRECT("\'"&D{r}&"\'!I12"),""),"")',
        )
        style_calc(c)
        style_input(ws.cell(row=r, column=12, value=""))

    total_r = 31
    ws.cell(row=total_r, column=1, value="").fill = FILL_HEADER
    ws.cell(row=total_r, column=2, value="TỔNG CỘNG").font = FONT_WHITE
    ws.cell(row=total_r, column=2).fill = FILL_HEADER
    for col in range(3, 5):
        ws.cell(row=total_r, column=col).fill = FILL_HEADER
    for col, letter in [(5, "E"), (6, "F"), (7, "G"), (8, "H"), (9, "I")]:
        c = ws.cell(row=total_r, column=col, value=f"=SUM({letter}5:{letter}29)")
        c.fill = FILL_HEADER
        c.font = FONT_WHITE
        c.number_format = VND
    for col in range(10, 13):
        ws.cell(row=total_r, column=col).fill = FILL_HEADER

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:L29"

    # Seed amounts for jobs without sheets (from TaiChinh data as contract value)
    manual_seed = {
        7: 10_700_000,  # row for Bizeyes Thảo - row 7 is stt 3 -> r=7
    }
    # rows: stt3 r=7 Rivus 10700000, stt4 r=8 Masteris 2500000, etc.
    seeds_money = [
        (7, 10_700_000),
        (8, 2_500_000),
        (9, 5_000_000),
        (10, 7_000_000),
        (11, 3_500_000),
        (12, 20_000_000),
        (13, 23_000_000),
        (14, 26_000_000),
    ]
    for r, val in seeds_money:
        # only if no sheet (D empty) — they are empty so set E (tổng HĐ)
        ws.cell(row=r, column=5).value = val
        money(ws.cell(row=r, column=5), is_input=True)
        # remaining formula already E-F


def build_cong_no(wb):
    ws = wb.create_sheet("03_CONG_NO", 3)
    set_col_widths(ws, [5, 22, 24, 16, 16, 16, 14, 36])
    add_title(ws, "03 · CÔNG NỢ KHÁCH HÀNG (tổng → Dashboard)", 1, "H")

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        "Nguồn chính: 02_JOB_AGENCY (tự đồng bộ). Bluescope là 1 dòng trong Agency + file booking riêng. "
        "Có thể bổ sung khoản ngoài job ở cuối bảng (vàng)."
    )
    ws["A2"].font = FONT_SMALL
    ws["A2"].fill = FILL_SOFT

    headers = [
        "#",
        "Agency",
        "Tên job",
        "Tổng HĐ",
        "Đã thu",
        "Còn lại",
        "Trạng thái thu",
        "Ghi chú",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 1, 8)

    # Pull 25 rows from Job Agency
    for i in range(25):
        r = 5 + i
        ar = 5 + i  # agency row
        ws.cell(row=r, column=1, value=i + 1).border = THIN
        for col, src in [
            (2, f"='02_JOB_AGENCY'!B{ar}"),
            (3, f"='02_JOB_AGENCY'!C{ar}"),
            (4, f"='02_JOB_AGENCY'!E{ar}"),
            (5, f"='02_JOB_AGENCY'!F{ar}"),
            (6, f"='02_JOB_AGENCY'!G{ar}"),
        ]:
            c = ws.cell(row=r, column=col, value=src)
            if col >= 4:
                money(c)
            else:
                style_calc(c)
        c = ws.cell(
            row=r,
            column=7,
            value=(
                f'=IF(OR(B{r}="",D{r}=0),"",'
                f'IF(E{r}>=D{r},"✅ Đã thu",'
                f'IF(E{r}>0,"⚠️ Một phần","🔴 Chưa thu")))'
            ),
        )
        style_calc(c)
        style_input(ws.cell(row=r, column=8, value=""))

    # Extra manual rows
    ws.cell(row=31, column=2, value="— Khoản ngoài job (nhập tay) —").font = FONT_BOLD
    for i in range(5):
        r = 32 + i
        ws.cell(row=r, column=1, value=26 + i).border = THIN
        for col in range(2, 4):
            style_input(ws.cell(row=r, column=col, value=""))
        money(ws.cell(row=r, column=4, value=0), is_input=True)
        money(ws.cell(row=r, column=5, value=0), is_input=True)
        c = ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
        money(c)
        c = ws.cell(
            row=r,
            column=7,
            value=(
                f'=IF(D{r}=0,"",IF(E{r}>=D{r},"✅ Đã thu",IF(E{r}>0,"⚠️ Một phần","🔴 Chưa thu")))'
            ),
        )
        style_calc(c)
        style_input(ws.cell(row=r, column=8, value=""))

    tr = 38
    ws.cell(row=tr, column=2, value="TỔNG CÔNG NỢ").font = FONT_WHITE
    ws.cell(row=tr, column=2).fill = FILL_HEADER
    for col, letter in [(4, "D"), (5, "E"), (6, "F")]:
        c = ws.cell(row=tr, column=col, value=f"=SUM({letter}5:{letter}36)")
        c.fill = FILL_HEADER
        c.font = FONT_WHITE
        c.number_format = VND
    for col in [1, 3, 7, 8]:
        ws.cell(row=tr, column=col).fill = FILL_HEADER

    ws["B40"] = "→ Dashboard đọc F38 (Tổng còn lại) và D38 (Tổng HĐ)."
    ws["B40"].font = FONT_SMALL
    ws.freeze_panes = "A5"


def build_no_ca_nhan(wb):
    ws = wb.create_sheet("04_NO_CA_NHAN", 4)
    set_col_widths(ws, [8, 14, 16, 14, 14, 16, 24])
    add_title(ws, "04 · NỢ CÁ NHÂN & LỊCH TRẢ", 1, "G")

    # MOM
    ws.merge_cells("A3:G3")
    ws["A3"] = "1. NỢ MOM (trả cố định / tháng)"
    ws["A3"].fill = FILL_SECTION
    ws["A3"].font = FONT_SECTION

    ws["A4"] = "Nợ gốc"
    ws["B4"] = "='99_CAU_HINH'!C7"
    money(ws["B4"])
    ws["C4"] = "Trả/tháng"
    ws["D4"] = "='99_CAU_HINH'!C8"
    money(ws["D4"])
    ws["E4"] = "Số tháng đã trả"
    ws["F4"] = 0
    money(ws["F4"], is_input=True)

    ws["A5"] = "Nợ MOM còn lại"
    ws["F5"] = "=B4-(D4*F4)"
    money(ws["F5"])
    ws["G5"] = "← Cập nhật F4 khi trả"
    ws["G5"].font = FONT_SMALL

    headers = ["Tháng", "Kỳ", "Số tiền trả", "Ngày trả", "Đã trả?", "Nợ còn lại", "Ghi chú"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=7, column=i, value=h)
    style_header_row(ws, 7, 1, 7)

    # 24 months from T7/2026
    months = []
    y, m = 2026, 7
    for i in range(1, 25):
        months.append((i, f"T{m}/{y}"))
        m += 1
        if m > 12:
            m = 1
            y += 1

    for i, (stt, ky) in enumerate(months):
        r = 8 + i
        ws.cell(row=r, column=1, value=stt).border = THIN
        ws.cell(row=r, column=2, value=ky).border = THIN
        c = ws.cell(row=r, column=3, value="=$D$4")
        money(c)
        style_input(ws.cell(row=r, column=4, value=""))
        style_input(ws.cell(row=r, column=5, value=""))
        if i == 0:
            c = ws.cell(row=r, column=6, value="=$B$4-C8")
        else:
            c = ws.cell(row=r, column=6, value=f"=F{r-1}-C{r}")
        money(c)
        style_input(ws.cell(row=r, column=7, value=""))

    # Trí
    r0 = 34
    ws.merge_cells(f"A{r0}:G{r0}")
    ws[f"A{r0}"] = "2. NỢ A TRÍ"
    ws[f"A{r0}"].fill = FILL_SECTION
    ws[f"A{r0}"].font = FONT_SECTION

    ws[f"A{r0+1}"] = "Nợ gốc"
    ws[f"B{r0+1}"] = "='99_CAU_HINH'!C9"
    money(ws[f"B{r0+1}"])
    ws[f"C{r0+1}"] = "Đã trả"
    ws[f"D{r0+1}"] = 0
    money(ws[f"D{r0+1}"], is_input=True)
    ws[f"E{r0+1}"] = "Còn lại"
    ws[f"F{r0+1}"] = f"=B{r0+1}-D{r0+1}"
    money(ws[f"F{r0+1}"])

    # Log trả Trí
    ws[f"A{r0+3}"] = "Lịch / log trả a Trí"
    ws[f"A{r0+3}"].font = FONT_BOLD
    for i, h in enumerate(["#", "Ngày", "Số tiền", "Hình thức", "Ghi chú", "", ""], 1):
        ws.cell(row=r0 + 4, column=i, value=h if h else None)
    style_header_row(ws, r0 + 4, 1, 5)
    for i in range(10):
        r = r0 + 5 + i
        ws.cell(row=r, column=1, value=i + 1).border = THIN
        for col in range(2, 6):
            if col == 3:
                money(ws.cell(row=r, column=col, value=0), is_input=True)
            else:
                style_input(ws.cell(row=r, column=col, value=""))

    ws[f"A{r0+16}"] = "Tổng đã trả (log)"
    ws[f"C{r0+16}"] = f"=SUM(C{r0+5}:C{r0+14})"
    money(ws[f"C{r0+16}"])
    ws[f"D{r0+16}"] = "Có thể copy số này sang D35"
    ws[f"D{r0+16}"].font = FONT_SMALL

    # Export for dashboard
    ws["I1"] = "EXPORT"
    ws["I2"] = "No_MOM"
    ws["J2"] = "=F5"
    money(ws["J2"])
    ws["I3"] = "No_Tri"
    ws["J3"] = "=F35"
    money(ws["J3"])


def build_tin_dung(wb):
    ws = wb.create_sheet("05_TIN_DUNG", 5)
    set_col_widths(ws, [6, 16, 12, 14] + [12] * 12)
    add_title(ws, "05 · TÍN DỤNG — Dư nợ thẻ theo tháng", 1, "P")

    headers = [
        "STT",
        "Ngân hàng",
        "Dòng",
        "Dư nợ gốc",
        "T7/26",
        "T8/26",
        "T9/26",
        "T10/26",
        "T11/26",
        "T12/26",
        "T1/27",
        "T2/27",
        "T3/27",
        "T4/27",
        "T5/27",
        "T6/27",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 1, 16)

    banks = [
        (1, "TPBANK", 2_000_000),
        (2, "TECHCOMBANK", 2_500_000),
        (3, "OCB", 2_000_000),
        (4, "SCB", 4_000_000),
    ]
    row = 4
    for stt, bank, du in banks:
        # Đã trả row
        ws.cell(row=row, column=1, value=stt).border = THIN
        ws.cell(row=row, column=2, value=bank).border = THIN
        ws.cell(row=row, column=3, value="💳 Đã trả").border = THIN
        money(ws.cell(row=row, column=4, value=du), is_input=True)
        for col in range(5, 17):
            money(ws.cell(row=row, column=col, value=0), is_input=True)
        # Còn lại row
        r2 = row + 1
        ws.cell(row=r2, column=1).border = THIN
        ws.cell(row=r2, column=2).border = THIN
        ws.cell(row=r2, column=3, value="✅ Còn lại").border = THIN
        money(ws.cell(row=r2, column=4, value=f"=D{row}"))
        # remaining cascade: first month = gốc - đã trả T7, then previous - paid
        for col in range(5, 17):
            letter = get_column_letter(col)
            if col == 5:
                formula = f"=D{r2}-{letter}{row}"
            else:
                prev = get_column_letter(col - 1)
                formula = f"={prev}{r2}-{letter}{row}"
            money(ws.cell(row=r2, column=col, value=formula))
        row += 2

    # Total remaining current (col D of còn lại rows or last month)
    ws.cell(row=row + 1, column=2, value="TỔNG DƯ NỢ HIỆN TẠI (cột gốc)").font = FONT_BOLD
    # sum of D for "Còn lại" rows: 5,7,9,11
    ws.cell(row=row + 1, column=4, value="=D5+D7+D9+D11")
    money(ws.cell(row=row + 1, column=4))

    ws["B" + str(row + 3)] = (
        "Nhập số ĐÃ TRẢ từng tháng (ô vàng). Dòng Còn lại tự trừ dồn. Dashboard đọc tổng dư nợ gốc."
    )
    ws["B" + str(row + 3)].font = FONT_SMALL

    # Export
    ws["R1"] = "EXPORT_Tong_TinDung"
    ws["S1"] = f"=D{row+1}"
    money(ws["S1"])


def build_dashboard(wb):
    ws = wb.create_sheet("01_DASHBOARD", 1)
    set_col_widths(ws, [4, 28, 18, 18, 18, 18, 18, 22])
    add_title(ws, "💰 MONEY 2026 — DASHBOARD CHÍNH", 1, "H")

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Bluescope booking (file riêng): {BLUESCOPE_URL}  |  "
        "Job Agency & công nợ tự tổng hợp từ các sheet JOB_*"
    )
    ws["A2"].font = FONT_SMALL
    ws["A2"].fill = FILL_LINK

    # A. Tổng quan tài chính
    ws.merge_cells("B4:D4")
    ws["B4"] = "A. TỔNG QUAN TÀI CHÍNH"
    ws["B4"].fill = FILL_SECTION
    ws["B4"].font = FONT_SECTION

    ws["B5"] = "Khoản mục"
    ws["C5"] = "Số tiền (đ)"
    ws["D5"] = "Ghi chú"
    style_header_row(ws, 5, 2, 4)

    kpis = [
        (6, "💵 Tiền hiện có", "='99_CAU_HINH'!C6", "Sửa ở 99_CAU_HINH"),
        (7, "🏦 Nợ MOM còn lại", "='04_NO_CA_NHAN'!J2", "Từ lịch trả MOM"),
        (8, "👤 Nợ a Trí còn lại", "='04_NO_CA_NHAN'!J3", "Từ log trả Trí"),
        (9, "💳 Tín dụng đang dùng", "='05_TIN_DUNG'!S1", "Tổng dư thẻ"),
        (10, "📊 Tổng nợ phải trả", "=C7+C8+C9", "MOM + Trí + Thẻ"),
        (11, "📈 Thực có (Tiền − Nợ)", "=C6-C10", "Âm = đang âm vốn"),
    ]
    for r, label, formula, note in kpis:
        ws.cell(row=r, column=2, value=label).font = FONT_BOLD
        c = ws.cell(row=r, column=3, value=formula)
        money(c)
        ws.cell(row=r, column=4, value=note).font = FONT_SMALL
        if r == 11:
            c.fill = FILL_ORANGE
            c.font = FONT_MONEY

    # B. Công nợ KH
    ws.merge_cells("B13:G13")
    ws["B13"] = "B. CÔNG NỢ KHÁCH HÀNG (từ 03_CONG_NO ← Job Agency ← job sheets + Bluescope)"
    ws["B13"].fill = FILL_SECTION
    ws["B13"].font = FONT_SECTION

    ws["B14"] = "Tổng giá trị HĐ / booking"
    ws["C14"] = "='03_CONG_NO'!D38"
    money(ws["C14"])
    ws["B15"] = "Đã thu"
    ws["C15"] = "='03_CONG_NO'!E38"
    money(ws["C15"])
    ws["B16"] = "🔴 TỔNG CÒN PHẢI THU"
    ws["C16"] = "='03_CONG_NO'!F38"
    money(ws["C16"])
    ws["C16"].fill = FILL_WARN
    ws["C16"].font = FONT_MONEY
    ws["D16"] = "← Đây là tổng công nợ liên kết Dashboard"
    ws["D16"].font = FONT_SMALL

    # C. P&L jobs
    ws.merge_cells("B18:G18")
    ws["B18"] = "C. TỔNG P&L THEO JOB AGENCY"
    ws["B18"].fill = FILL_SECTION
    ws["B18"].font = FONT_SECTION

    ws["B19"] = "Tổng chi phí jobs"
    ws["C19"] = "='02_JOB_AGENCY'!H31"
    money(ws["C19"])
    ws["B20"] = "Tổng lợi nhuận gộp"
    ws["C20"] = "='02_JOB_AGENCY'!I31"
    money(ws["C20"])
    ws["C20"].fill = FILL_GREEN

    # Mini table top jobs
    ws["B22"] = "#"
    ws["C22"] = "Agency / Job"
    ws["D22"] = "Tổng HĐ"
    ws["E22"] = "Còn lại"
    ws["F22"] = "Lợi nhuận"
    ws["G22"] = "Trạng thái"
    style_header_row(ws, 22, 2, 7)

    for i in range(10):
        r = 23 + i
        ar = 5 + i
        ws.cell(row=r, column=2, value=i + 1).border = THIN
        c = ws.cell(
            row=r,
            column=3,
            value=f"=IF('02_JOB_AGENCY'!B{ar}=\"\",\"\",'02_JOB_AGENCY'!B{ar}&\" — \"&'02_JOB_AGENCY'!C{ar})",
        )
        style_calc(c)
        for col, letter in [(4, "E"), (5, "G"), (6, "I")]:
            c = ws.cell(
                row=r,
                column=col,
                value=f"=IF('02_JOB_AGENCY'!B{ar}=\"\",\"\",'02_JOB_AGENCY'!{letter}{ar})",
            )
            money(c)
        c = ws.cell(
            row=r,
            column=7,
            value=f"=IF('02_JOB_AGENCY'!B{ar}=\"\",\"\",'02_JOB_AGENCY'!J{ar})",
        )
        style_calc(c)

    # D. Quick links
    ws.merge_cells("B34:G34")
    ws["B34"] = "D. LIÊN KẾT NHANH"
    ws["B34"].fill = FILL_SECTION
    ws["B34"].font = FONT_SECTION

    links = [
        (35, "Job Agency (sổ cái)", "Xem sheet 02_JOB_AGENCY"),
        (36, "Công nợ chi tiết", "Xem sheet 03_CONG_NO"),
        (37, "Job LG (mẫu)", "Xem sheet JOB_LG — copy _TEMPLATE_JOB để thêm job"),
        (38, "Job Bluescope", "Xem sheet JOB_BLUESCOPE"),
        (39, "Mở Bluescope Booking (file riêng)", BLUESCOPE_URL),
        (40, "Cấu hình % / tiền hiện có", "Xem sheet 99_CAU_HINH"),
    ]
    for r, label, val in links:
        ws.cell(row=r, column=2, value=label).font = FONT_BOLD
        cell = ws.cell(row=r, column=3, value=val)
        if str(val).startswith("http"):
            cell.hyperlink = val
            cell.font = FONT_LINK
            cell.fill = FILL_LINK
        else:
            cell.font = FONT_SMALL

    ws["B42"] = (
        "Net sức khỏe = Tiền hiện có + Công nợ phải thu − Tổng nợ phải trả  →  "
    )
    ws["B42"].font = FONT_BOLD
    ws["C42"] = "=C6+C16-C10"
    money(ws["C42"])
    ws["C42"].fill = FILL_ORANGE
    ws["D42"] = "Tiền + phải thu − phải trả"
    ws["D42"].font = FONT_SMALL

    ws.freeze_panes = "A4"


def build():
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    build_huong_dan(wb)
    build_cau_hinh(wb)
    build_dashboard(wb)
    build_job_agency(wb)
    build_cong_no(wb)
    build_no_ca_nhan(wb)
    build_tin_dung(wb)
    build_job_template(wb, "_TEMPLATE_JOB")
    build_job_lg(wb)
    build_job_bluescope(wb)

    # Reorder sheets nicely
    order = [
        "00_HUONG_DAN",
        "01_DASHBOARD",
        "02_JOB_AGENCY",
        "03_CONG_NO",
        "04_NO_CA_NHAN",
        "05_TIN_DUNG",
        "JOB_LG",
        "JOB_BLUESCOPE",
        "_TEMPLATE_JOB",
        "99_CAU_HINH",
    ]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(OUT)
    print(f"Wrote {OUT}")
    return OUT


if __name__ == "__main__":
    build()
